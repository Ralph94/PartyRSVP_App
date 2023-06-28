from openpyxl import Workbook

# Create a new workbook
workbook = Workbook()

# Select the active sheet
sheet = workbook.active

# Data source (example list of dictionaries)
data = [
    {'Name': 'John Doe', 'Age': 30},
    {'Name': 'Jane Smith', 'Age': 25},
    {'Name': 'Michael Johnson', 'Age': 35}
]

# Write headers
headers = list(data[0].keys())
for col_num, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col_num).value = header

# Write data rows
for row_num, row_data in enumerate(data, start=2):
    for col_num, value in enumerate(row_data.values(), start=1):
        sheet.cell(row=row_num, column=col_num).value = value

# Save the workbook
workbook.save('Book1.xlsx')

